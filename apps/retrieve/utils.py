#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = 'heianhu'
import zipfile
from crawl_data.models import Summary, Detail, Authors, Periodicals, References, ReferencesCBBD, ReferencesCCND, \
    ReferencesCDFD, ReferencesCJFQ, ReferencesCPFD, ReferencesCMFD, ReferencesCRLDENG, ReferencesSSJD, Organization
from openpyxl import Workbook
import time
from ZhiWang.settings import BASE_DIR


def write_to_txt(getdetailinfo_id):
    """
    将指定id的文章写入txt
    :param getdetailinfo_id: 文章summary中的id号
    :return: 生成的excel文件名(文章detail中的detail_id),失败时返回None
    """
    summary_id = int(getdetailinfo_id)
    # 在Excel中插入数据，文件名为detail_id字段

    fields = ['FN', 'VR', 'PT', 'AU', 'AF', 'BA', 'CA', 'GP', 'BE', 'TI', 'SO', 'SE', 'BS', 'LA', 'DT',
              'CT', 'CY', 'CL', 'SP', 'HO', 'DE', 'ID', 'AB', 'C1', 'RP', 'EM', 'FU', 'FX', 'CR', 'NR',
              'TC', 'Z9', 'PU', 'PI', 'PA', 'SN', 'BN', 'J9', 'JI', 'PD', 'PY', 'VL', 'IS', 'SI', 'PN',
              'SU', 'BP', 'EP', 'AR', 'DI', 'D2', 'PG', 'P2', 'WC', 'SC', 'GA', 'UT', 'ER', 'EF']
    values = {field: [i, ] for i, field in enumerate(fields)}
    article_summary = Summary.objects.get(id=summary_id)
    try:
        article_detail = Detail.objects.get(id=article_summary.detail_id)
    except Exception as e:
        print('部分summary没有detail', e)
        return None

    # 文件名
    values['FN'].append(article_detail.detail_id)
    # 作者
    # values['AU'][1] = article_summary.authors
    for author in article_summary.detail.authors.split():
        if author.isdigit():
            values['AU'].append(Authors.objects.filter(id=int(author))[0].authors_name)
        else:
            values['AU'].append(author)
    # 标题
    values['TI'].append(article_summary.title)
    # 出版日期
    values['PD'].append(article_summary.issuing_time.strftime('%Y/%m/%d'))
    # issn号
    values['SN'].append(Periodicals.objects.get(id=article_summary.source_id).issn_number)
    # 出版物名称
    values['SO'].append(Periodicals.objects.get(id=article_summary.source_id).name)
    # 摘要
    values['AB'].append(article_detail.detail_abstract)

    try:
        # 参考文献
        article_reference = References.objects.get(id=article_detail.references_id)
    except References.DoesNotExist:
        # 处理没有文献情况
        values['CR'].append('')

    else:
        str_refer = ['CBBD', 'CDFD', 'CJFQ', 'CMFD', 'CRLDENG', 'SSJD', 'CCND', 'CPFD']
        refers = []  # 参考于各个期刊的id，每个期刊的id以列表形式存在 [[1,2],[3,4]]
        for refer in str_refer:
            refers.append(getattr(article_reference, refer).split())

        all_refers = [
            ReferencesCBBD.objects.filter(id__in=refers[0]),
            ReferencesCDFD.objects.filter(id__in=refers[1]),
            ReferencesCJFQ.objects.filter(id__in=refers[2]),
            ReferencesCMFD.objects.filter(id__in=refers[3]),
            ReferencesCRLDENG.objects.filter(id__in=refers[4]),
            ReferencesSSJD.objects.filter(id__in=refers[5]),
            ReferencesCCND.objects.filter(id__in=refers[6]),
            ReferencesCPFD.objects.filter(id__in=refers[7]),
        ]            # 每一项是相应期刊的查询集

        # # 对每一条结果取title值
        # all_refers_title = (refer.title for queryset in all_refers for refer in queryset)
        # # 将每个title分别填入
        # for refers_title in all_refers_title:
        #     values['CR'].append(refers_title)

        # 对每一条结果逐条获取数据
        all_refers = (refer for queryset in all_refers for refer in queryset)
        # 将每个title分别填入
        for refers in all_refers:
            try:
                temp_refers_info = (refers.title, refers.authors, refers.source, refers.issuing_time)
            except AttributeError:
                temp_refers_info = (refers.title, refers.info, refers.issuing_time)
            values['CR'].append(temp_refers_info)

    filename = '{0}.txt'.format(article_detail.detail_id)
    with open(BASE_DIR + '/media/txt/single/' + filename, 'w+', encoding='utf-8') as file:
        for key, (i, *all_value) in values.items():
            file.write(str(key))
            for num, v in enumerate(all_value):
                file.write('    ' + ''.join(v))
                if num != len(all_value)-1:
                    file.write('\n')
            file.write('\n')

    return filename


def compress_txt(ids):
    """
    将批量id的文章写入txt，然后压缩成一个zip文件
    :param ids: summary文章id列表
    :return: 生成的zip文件名(不包含路径)
    """
    files = []
    zip_name = time.strftime("%Y%m%d%H%M%S", time.localtime()) + '.zip'
    jungle_zip = zipfile.ZipFile(BASE_DIR + '/media/txt/{0}'.format(zip_name), 'w')

    for id in ids:
        filename = write_to_txt(id)
        file = BASE_DIR + '/media/txt/single/{0}'.format(filename)
        jungle_zip.write(file, filename, compress_type=zipfile.ZIP_DEFLATED)

    jungle_zip.close()

    return zip_name



def write_to_excel(getdetailinfo_id):
    """
    将指定id的文章写入excel
    :param getdetailinfo_id:文章summary中的id号
    :return: 生成的excel文件名(文章detail中的detail_id),失败时返回None
    """
    summary_id = int(getdetailinfo_id)
    # 在Excel中插入数据，文件名为detail_id字段

    fields = ['FN', 'VR', 'PT', 'AU', 'AF', 'BA', 'CA', 'GP', 'BE', 'TI', 'SO', 'SE', 'BS', 'LA', 'DT',
              'CT', 'CY', 'CL', 'SP', 'HO', 'DE', 'ID', 'AB', 'C1', 'RP', 'EM', 'FU', 'FX', 'CR', 'NR',
              'TC', 'Z9', 'PU', 'PI', 'PA', 'SN', 'BN', 'J9', 'JI', 'PD', 'PY', 'VL', 'IS', 'SI', 'PN',
              'SU', 'BP', 'EP', 'AR', 'DI', 'D2', 'PG', 'P2', 'WC', 'SC', 'GA', 'UT', 'ER', 'EF']

    values = {field: [i, ] for i, field in enumerate(fields)}
    article_summary = Summary.objects.get(id=summary_id)
    try:
        article_detail = Detail.objects.get(id=article_summary.detail_id)
    except Exception as e:
        print('部分summary没有detail', e)
        return None

    # 文件名
    values['FN'].append(article_detail.detail_id)
    # 作者
    # values['AU'][1] = article_summary.authors
    for author in article_summary.detail.authors.split():
        if author.isdigit():
            values['AU'].append(Authors.objects.filter(id=int(author))[0].authors_name)
        else:
            values['AU'].append(author)
    # 标题
    values['TI'].append(article_summary.title)
    # 出版日期
    values['PD'].append(article_summary.issuing_time.strftime('%Y/%m/%d'))
    # issn号
    values['SN'].append(Periodicals.objects.get(id=article_summary.source_id).issn_number)
    # 出版物名称
    values['SO'].append(Periodicals.objects.get(id=article_summary.source_id).name)
    # 摘要
    values['AB'].append(article_detail.detail_abstract)

    try:
        # 参考文献
        article_reference = References.objects.get(id=article_detail.references_id)
    except References.DoesNotExist:
        # 处理没有文献情况
        values['CR'].append('')

    else:
        str_refer = ['CBBD', 'CDFD', 'CJFQ', 'CMFD', 'CRLDENG', 'SSJD', 'CCND', 'CPFD']
        refers = []  # 参考于各个期刊的id，每个期刊的id以列表形式存在 [[1,2],[3,4]]
        for refer in str_refer:
            refers.append(getattr(article_reference, refer).split())

        all_refers = []  # 每一项是相应期刊的查询集
        all_refers.append(ReferencesCBBD.objects.filter(id__in=refers[0]))
        all_refers.append(ReferencesCDFD.objects.filter(id__in=refers[1]))
        all_refers.append(ReferencesCJFQ.objects.filter(id__in=refers[2]))
        all_refers.append(ReferencesCMFD.objects.filter(id__in=refers[3]))
        all_refers.append(ReferencesCRLDENG.objects.filter(id__in=refers[4]))
        all_refers.append(ReferencesSSJD.objects.filter(id__in=refers[5]))
        all_refers.append(ReferencesCCND.objects.filter(id__in=refers[6]))
        all_refers.append(ReferencesCPFD.objects.filter(id__in=refers[7]))

        # # 对每一条结果取title值
        # all_refers_title = (refer.title for queryset in all_refers for refer in queryset)
        # # 将每个title分别填入
        # for refers_title in all_refers_title:
        #     values['CR'].append(refers_title)

        # 对每一条结果逐条获取数据
        all_refers = (refer for queryset in all_refers for refer in queryset)
        # 将每个title分别填入
        for refers in all_refers:
            try:
                temp_refers_info = (refers.title, refers.authors, refers.source, refers.issuing_time)
            except AttributeError:
                temp_refers_info = (refers.title, refers.info, refers.issuing_time)
            values['CR'].append(temp_refers_info)

    wb = Workbook()
    sheet = wb.get_active_sheet()

    # for k, [i, v] in values.items():
    #     sheet.cell(row=i + 1, column=1).value = k
    #     sheet.cell(row=i + 1, column=2).value = v
    for key, (i, *all_value) in values.items():
        sheet.cell(row=i + 1, column=1).value = key
        for _, v in enumerate(all_value):
            sheet.cell(row=i + 1, column=2 + _).value = ''.join(v)

    # 保存并返回下载
    # filename = '{0}.xlsx'.format(article_detail.detail_id)
    filename = '{0}.xlsx'.format(article_detail.detail_id)

    wb.save(BASE_DIR + '/media/excel/single/' + filename)
    return filename


def compress_excel(ids):
    """
    将批量id的文章写入excel，然后压缩成一个zip文件
    :param ids: summary文章id列表
    :return: 生成的zip文件名(不包含路径)
    """
    files = []
    for id in ids:
        filename = write_to_excel(id)
        files.append(BASE_DIR + '/media/excel/single/{0}'.format(filename))

    timestamp = str(time.time()).replace('.', '')
    zip_name = '{0}.zip'.format(timestamp)
    jungle_zip = zipfile.ZipFile(BASE_DIR + '/media/excel/{0}'.format(zip_name), 'w')
    for file in files:
        jungle_zip.write(file, compress_type=zipfile.ZIP_DEFLATED)
    jungle_zip.close()

    return zip_name
