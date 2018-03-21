from django.shortcuts import render, redirect, render_to_response
from django.views.generic import View
from retrieve.models import SearchFilter
from crawl_data.models import Summary, Detail, Authors, Periodicals, References, ReferencesCBBD, ReferencesCCND, \
    ReferencesCDFD, ReferencesCJFQ, ReferencesCPFD, ReferencesCMFD, ReferencesCRLDENG, ReferencesSSJD, Organization
from django.db.models import Q
import jieba
import jieba.posseg
from pure_pagination import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse, FileResponse, JsonResponse, HttpResponseRedirect
from django.utils.http import urlquote
from django.http import Http404
from openpyxl import Workbook
import json
import time
import datetime


# Create your views here.

class IndexView(View):
    def get(self, request):
        if not request.session.session_key:
            request.session.save()
        # try:
        #     request.session.save()
        # except:
        #     pass
        return render(request, 'index.html', {
        })


class Search(View):
    def post(self, request):
        """
            'txt_2_sel': $('#txt_2_sel').val(),{# 主题选择 #}
            'txt_2_value1': $('#txt_2_value1').val(),{# 主题输入框1 #}
            'txt_2_relation':  $('#txt_2_relation').val(),{# 并行条件选择 #}
            'txt_2_value2': $('#txt_2_value2').val(),{# 主题输入框2 #}
            'au_1_sel': $('#au_1_sel').val(), {# 作者选择 #}
            'au_1_value1': $('#au_1_value1').val(),   {# 作者输入框 #}
            'magazine_value1': $('#magazine_value1').val(),{# 文献来源输入框 #}
        :param request:
        :return:
        """

        para = dict(
            txt_2_sel=request.POST.get('txt_2_sel', ''),  # 主题选择
            txt_2_value1=request.POST.get('txt_2_value1', ''),  # 主题输入框1
            txt_2_relation=request.POST.get('txt_2_relation', ''),  # 并行条件选择
            txt_2_value2=request.POST.get('txt_2_value2', ''),  # 主题输入框2
            # au_1_sel = request.POST.get('au_1_sel', '') # 作者选择  暂时没用
            au_1_value1=request.POST.get('au_1_value1', ''),  # 作者输入框
            au_special=request.POST.get('au_special', ''),  # 作者模糊/精准
            org_1_value=request.POST.get('org_1_value', ''),  # 组织
            org_1_special2=request.POST.get('org_1_special2', ''),  # 组织模糊/精准
            publishdate_from=request.POST.get('publishdate_from', ''),  # 起始年
            publishdate_to=request.POST.get('publishdate_to', ''),  # 截止年
            magazine_value1=request.POST.get('magazine_value1', ''),  # 文献来源输入框
            magazine_special=request.POST.get('magazine_special', '')  # 文献模糊/精准
        )

        search_filter = SearchFilter()
        search_filter.session_id = request.session.session_key
        # queryId = datetime.now()
        queryId = str(time.time()).replace('.', '')
        search_filter.time = queryId
        search_filter.filterPara = str(para)
        search_filter.save()
        return redirect('/retrieve?queryId={}'.format(queryId))

    def get(self, request):
        """
        搜索功能
        目前能根据标题和作者，在被标记的期刊中的文献进行搜索
        :param request:
        :return:
        """
        # 获取关键词和搜索类型
        queryId = request.GET.get('queryId', '')
        curr_page = request.GET.get('page', '1')
        session_key = request.session.session_key
        try:
            search_filter = SearchFilter.objects.get(time=queryId, session_id=session_key)
        except KeyError:
            return render(request, 'index.html')

        search_filter = eval(search_filter.filterPara)

        try:
            txt_2_sel = search_filter.get('txt_2_sel', '')
            txt_2_value1 = search_filter.get('txt_2_value1', '')
            txt_2_relation = search_filter.get('txt_2_relation', '')
            txt_2_value2 = search_filter.get('txt_2_value2', '')
            au_1_value1 = search_filter.get('au_1_value1', '')
            org_1_value = search_filter.get('org_1_value', '')
            publishdate_from = search_filter.get('publishdate_from', '')
            publishdate_to = search_filter.get('publishdate_to', '')
            magazine_value1 = search_filter.get('magazine_value1', '')

            txt_2_sel_dic = {  # CNKI_AND CNKI_OR
                "SU": (Q(title__icontains=txt_2_value1), Q(title__icontains=txt_2_value2)),  # 标题
                "KY": (Q(detail_keywords__icontains=txt_2_value1), Q(detail_keywords__icontains=txt_2_value2)),  # 关键词
                "AB": (Q(detail_abstract__icontains=txt_2_value1), Q(detail_abstract__icontains=txt_2_value2)),  # 摘要
                "CLC$=|?": (Q(issn_number=txt_2_value1), Q(issn_number=txt_2_value1))  # 中图分类号
            }
            org_id = Organization.objects.filter(organization_name__icontains=org_1_value)[0]
            org_id = str(org_id.id)

            if publishdate_from and publishdate_to:
                publishdate_from = publishdate_from.split('-')
                date_from = datetime.datetime(int(publishdate_from[0]), int(publishdate_from[1]),
                                              int(publishdate_from[2]), 0, 0)
                publishdate_to = publishdate_to.split('-')
                date_to = datetime.datetime(int(publishdate_to[0]), int(publishdate_to[1]), int(publishdate_to[2]),
                                            0, 0)
                date_filter = Q(issuing_time__range=(date_from, date_to))
            else:
                date_filter = Q()

            else_sel = Q(authors__icontains=au_1_value1) & \
                       (Q(source__issn_number__icontains=magazine_value1) | Q(source__name__icontains=magazine_value1)) \
                       & (Q(detail__organizations__icontains=org_1_value) | Q(detail__organizations__icontains=org_id)) \
                       & date_filter

            if txt_2_relation == 'CNKI_AND':
                all_articles = Summary.objects.filter(
                    (txt_2_sel_dic[txt_2_sel][0] & txt_2_sel_dic[txt_2_sel][1]) & else_sel)
            elif txt_2_relation == 'CNKI_OR':
                all_articles = Summary.objects.filter(
                    (txt_2_sel_dic[txt_2_sel][0] | txt_2_sel_dic[txt_2_sel][1]) & else_sel)
            elif txt_2_relation == 'CNKI_NOT':
                all_articles = Summary.objects.filter(
                    txt_2_sel_dic[txt_2_sel][0] & else_sel
                ).exclude(txt_2_sel_dic[txt_2_sel][1])
            else:
                return render(request, 'index.html')

            result_count = all_articles.count()
            # result_count = 1
            # all_articles = all_articles[start:end]


            # data = list(all_articles.values())

            # 分页功能
            try:
                page = request.GET.get('page', 1)
            except PageNotAnInteger:
                page = 1
            p = Paginator(all_articles, 12, request=request)
            articles = p.page(page)

            return render(request, 'index.html', {
                'all_articles': articles,
                # 'search_type': search_type,
                'keywords': txt_2_value1,
                'result_count': result_count
            })
        except KeyError:
            return render(request, 'index.html')


def write_to_excel(getdetailinfo_id):
    summary_id = int(getdetailinfo_id)
    # 在Excel中插入数据，文件名为detail_id字段

    fields = ['FN', 'VR', 'PT', 'AU', 'AF', 'BA', 'CA', 'GP', 'BE', 'TI', 'SO', 'SE', 'BS', 'LA', 'DT',
              'CT', 'CY', 'CL', 'SP', 'HO', 'DE', 'ID', 'AB', 'C1', 'RP', 'EM', 'FU', 'FX', 'CR', 'NR',
              'TC', 'Z9', 'PU', 'PI', 'PA', 'SN', 'BN', 'J9', 'JI', 'PD', 'PY', 'VL', 'IS', 'SI', 'PN',
              'SU', 'BP', 'EP', 'AR', 'DI', 'D2', 'PG', 'P2', 'WC', 'SC', 'GA', 'UT', 'ER', 'EF']

    values = {field: [i, ] for i, field in enumerate(fields)}
    article_summary = Summary.objects.get(id=summary_id)
    try:
        article_detail = Detail.objects.get(id=article_summary.detail_id)
    except Exception as e:
        print ('部分summary没有detail',e)
        return None

    # 文件名
    values['FN'].append(article_detail.detail_id)
    # 作者
    # values['AU'][1] = article_summary.authors
    for author in article_summary.detail.authors.split():
        if author.isdigit():
            values['AU'].append(Authors.objects.filter(id=int(author))[0].authors_name)
        else:
            values['AU'].append(author)
    # 标题
    values['TI'].append(article_summary.title)
    # 出版日期
    values['PD'].append(article_summary.issuing_time.strftime('%Y/%m/%d'))
    # issn号
    values['SN'].append(Periodicals.objects.get(id=article_summary.source_id).issn_number)
    # 出版物名称
    values['SO'].append(Periodicals.objects.get(id=article_summary.source_id).name)
    # 摘要
    values['AB'].append(article_detail.detail_abstract)

    try:
        # 参考文献
        article_reference = References.objects.get(id=article_detail.references_id)
    except References.DoesNotExist:
        # 处理没有文献情况
        values['CR'].append('')

    else:
        str_refer = ['CBBD', 'CDFD', 'CJFQ', 'CMFD', 'CRLDENG', 'SSJD', 'CCND', 'CPFD']
        refers = []  # 参考于各个期刊的id，每个期刊的id以列表形式存在 [[1,2],[3,4]]
        for refer in str_refer:
            refers.append(getattr(article_reference, refer).split())

        all_refers = []  # 每一项是相应期刊的查询集
        all_refers.append(ReferencesCBBD.objects.filter(id__in=refers[0]))
        all_refers.append(ReferencesCDFD.objects.filter(id__in=refers[1]))
        all_refers.append(ReferencesCJFQ.objects.filter(id__in=refers[2]))
        all_refers.append(ReferencesCMFD.objects.filter(id__in=refers[3]))
        all_refers.append(ReferencesCRLDENG.objects.filter(id__in=refers[4]))
        all_refers.append(ReferencesSSJD.objects.filter(id__in=refers[5]))
        all_refers.append(ReferencesCCND.objects.filter(id__in=refers[6]))
        all_refers.append(ReferencesCPFD.objects.filter(id__in=refers[7]))

        # # 对每一条结果取title值
        # all_refers_title = (refer.title for queryset in all_refers for refer in queryset)
        # # 将每个title分别填入
        # for refers_title in all_refers_title:
        #     values['CR'].append(refers_title)

        # 对每一条结果逐条获取数据
        all_refers = (refer for queryset in all_refers for refer in queryset)
        # 将每个title分别填入
        for refers in all_refers:
            try:
                temp_refers_info = (refers.title, refers.authors, refers.source, refers.issuing_time)
            except AttributeError:
                temp_refers_info = (refers.title, refers.info, refers.issuing_time)
            values['CR'].append(temp_refers_info)

    wb = Workbook()
    sheet = wb.get_active_sheet()

    # for k, [i, v] in values.items():
    #     sheet.cell(row=i + 1, column=1).value = k
    #     sheet.cell(row=i + 1, column=2).value = v
    for key, (i, *all_value) in values.items():
        sheet.cell(row=i + 1, column=1).value = key
        for _, v in enumerate(all_value):
            sheet.cell(row=i + 1, column=2 + _).value = ''.join(v)

    # 保存并返回下载
    filename = '{0}.xlsx'.format(article_detail.detail_id)
    wb.save('media/' + filename)
    return filename


class GetDetailInfo(View):

    def get(self, request, getdetailinfo_id):

        filename = write_to_excel(getdetailinfo_id)
        if not filename:
            response = render_to_response('404.html', {})
            response.status_code = 404
            return response
        file = open('media/{0}'.format(filename), 'rb')
        response = FileResponse(file)

        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{}"'.format(urlquote(filename))

        return response


class DownloadSel(View):

    def post(self, request):
        ids = request.POST.get('ids', '')
        ids = [int(id) for id in ids.split(',') if id]
        files = []
        for id in ids:
            filename = write_to_excel(id)
            files.append('media/{0}'.format(filename))

        import zipfile
        timestamp = str(time.time()).replace('.','')
        zip_name = '{0}.zip'.format(timestamp)
        jungle_zip = zipfile.ZipFile('media/{0}'.format(zip_name), 'w')
        for file in files:
            jungle_zip.write(file, compress_type=zipfile.ZIP_DEFLATED)
        jungle_zip.close()


        return JsonResponse({'status': 'success',
                         'data': zip_name.replace('.zip',''),
                         }, content_type='application/json')

class DownloadZip(View):
    def get(self, request, zip_name):
        zip_name = 'media/{0}.zip'.format(zip_name)
        file = open(zip_name, 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{}"'.format(urlquote(zip_name))

        return response


